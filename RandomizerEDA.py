from openpyxl import Workbook
from datetime import datetime
from sys import exit
from os import system
from random import uniform # returns floating-point value, use randint for integer value
XLSX_Worker = Workbook()
Active_XLSX = XLSX_Worker.active
Datetime_Create = datetime.now()

class Randomize_App:
    def __init__(self, start_int, end_int, result, attempts):
        self.start_int = start_int
        self.end_int = end_int
        self.result = result
        self.attempts = attempts

    def Welcome_Intro(self):
        system("CLS")
        print("Hello and Welcome To Randomize Cutout Generator")
        print("Exclusively Created for EDA which uses bond paper to do data.")
        print("Data will be automatically added to Excel... Choice the data at your own sake.\n")
        Active_XLSX.title = input('Please enter your name (Full Name) -> ')
        return

    def Intro_DataGet(self):
        print("Data Needed is 30 Attempts of Bond Paper Cutting")
        print("Short Bond Paper Dimension is 8.5 x 11 which this concludes the possible range to cut from 2.54cm to 27.94")
        print("Middle Cut is 13.97cm which should give a hint on where to start...")
        print("Please specify range cm to generate...")
        try:
            self.start_int = float(input("Input Starting Range in CM -> "))
            self.end_int = float(input("Input Starting Range in CM -> "))
            self.attempts = int(input("Input Retries, Will Represent as the Nth Column - > "))
            print("Generating Data...\n")
            return
        except ValueError:
            self.start_int = 0
            self.end_int = 0
            print("Error, you just inputted a value that is not a integer...")
            print("Rerun the program again... Exiting...")
            exit(0)

    def Randomize(self):
        ColumnCount = 0
        #for Column_nth in Active_XLSX.iter_cols(min_col = 1, max_col=self.attempts, min_row=1, max_row=20):
        for Column_nth in range(1, self.attempts + 1):
            ColumnCount += 1
            for Row_nth in range(1, 31):
                self.result = round(uniform(self.start_int, self.end_int), 1)
                Active_XLSX.cell(column = ColumnCount, row = Row_nth, value = (str(self.result) + ' cm'))
                print('Column #',Column_nth,'[ Column Position -> ',ColumnCount,'], Returned Output {}'.format(self.result), 'cm @ ', Row_nth)
        XLSX_Worker.save(Active_XLSX.title + '.xlsx')
            
    def XLSX_CreateWorksheet(self): # unused instance
        XLSX_Worker.create_sheet(Datetime_Create.isoformat())

Instance_1 = Randomize_App(0, 0, 0, 0)
#Instance_1.XLSX_CreateWorksheet()
Instance_1.Welcome_Intro()
Instance_1.Intro_DataGet()
Instance_1.Randomize()

print("Data Finished... Check", Active_XLSX.title + '.xlsx On the location of this script... Thank you!')
